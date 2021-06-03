"""
Advanced Python Project
"""

import openpyxl
import string
from openpyxl import Workbook
from openpyxl.styles import Font

class CheckPs:
    try:
        def __init__(self, n_1):
            self.n = n_1
    except Exception as e:
        print(e)

    try:
        def check_ps(self):
            path = "data.xlsx"
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj["Sheet1"]
            max_r = sheet_obj.max_row
            c = 0
            for i in range(1, max_r + 1):
                cell_obj = sheet_obj.cell(row=i, column=1)
                if cell_obj.value == self.n:
                    c = 1
            return c
    except Exception as e:
        print(e)

class PrintPs:
    try:
        @staticmethod
        def print_ps():
            path = "data.xlsx"
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj["Sheet1"]
            max_r = sheet_obj.max_row
            print("------ Welcome to The Python Project --------")
            print("The List of Available Ps ID :")
            for i in range(2, max_r + 1):
                cell_obj = sheet_obj.cell(row=i, column=1)
                print(cell_obj.value)
    except Exception as e:
        print(e)


class ExactPs:
    try:
        def __init__(self, n_1):
            self.n = n_1
    except Exception as e:
        print(e)
    try:
        def exactps(self):
            path = "data.xlsx"
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj["Sheet2"]
            max_r = sheet_obj.max_row
            for i in range(1, max_r + 1):
                cell_obj = sheet_obj.cell(row=i, column=1)
                if cell_obj.value == self.n:
                    return i
    except Exception as e:
        print(e)

class Detailsps:
    try:
        def __init__(self, n):
            self.n = n
    except Exception as e:
        print(e)
    try:
        def detailsPs(self):
            path = "data.xlsx"
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj_1 = wb_obj["Sheet1"]
            max_c = sheet_obj_1.max_column
            # out put file creation
            wb = Workbook()
            filepath = "output.xlsx"
            wb.save(filepath)
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active
            for letter in string.ascii_uppercase:
                sheet.column_dimensions[letter].width = 20
            for j in range(2, max_c + 1):
                cell_obj_1 = sheet_obj_1.cell(row=1, column=j)
                sheet.cell(row=j, column=1).value = cell_obj_1.value
                sheet.cell(row=j, column=1).font = Font(size=11, bold=True)
                cell_obj_2 = sheet_obj_1.cell(row=self.n, column=j)
                sheet.cell(row=j, column=2).value = cell_obj_2.value
                str = "The {c1} Grade is {c2}".format(c1=cell_obj_1.value, c2=cell_obj_2.value)
                print(str)
            print("---------------------------------------")
            sheet_obj_2 = wb_obj["Sheet2"]
            max_c = sheet_obj_2.max_column
            for j in range(2, max_c + 1):
                cell_obj_1 = sheet_obj_2.cell(row=1, column=j)
                sheet.cell(row=j, column=3).value = cell_obj_1.value
                sheet.cell(row=j, column=3).font = Font(size=11, bold=True)
                cell_obj_2 = sheet_obj_2.cell(row=self.n, column=j)
                sheet.cell(row=j, column=4).value = cell_obj_2.value
                str = "The {c1}  is {c2}".format(c1=cell_obj_1.value, c2=cell_obj_2.value)
                print(str)
            print("---------------------------------------")
            sheet_obj_3 = wb_obj["Sheet3"]
            max_c = sheet_obj_3.max_column
            for j in range(2, max_c + 1):
                cell_obj_1 = sheet_obj_3.cell(row=1, column=j)
                sheet.cell(row=j, column=5).value = cell_obj_1.value
                sheet.cell(row=j, column=5).font = Font(size=11, bold=True)
                cell_obj_2 = sheet_obj_3.cell(row=self.n, column=j)
                sheet.cell(row=j, column=6).value = cell_obj_2.value
                str = "The {c1}  person travelled is {c2}".format(c1=cell_obj_1.value, c2=cell_obj_2.value)
                print(str)
            print("---------------------------------------")
            sheet_obj_4 = wb_obj["Sheet4"]
            max_c = sheet_obj_4.max_column
            for j in range(2, max_c + 1):
                cell_obj_1 = sheet_obj_4.cell(row=1, column=j)
                sheet.cell(row=j, column=7).value = cell_obj_1.value
                sheet.cell(row=j, column=7).font = Font(size=11, bold=True)
                cell_obj_2 = sheet_obj_4.cell(row=self.n, column=j)
                sheet.cell(row=j, column=8).value = cell_obj_2.value
                str = "Programming Language {c1} has Expertise level is {c2}".format(c1=cell_obj_1.value,
                                                                                     c2=cell_obj_2.value)
                print(str)
            print("---------------------------------------")
            sheet_obj_5 = wb_obj["Sheet5"]
            max_c = sheet_obj_5.max_column
            for j in range(2, max_c + 1):
                cell_obj_1 = sheet_obj_5.cell(row=1, column=j)
                sheet.cell(row=j, column=9).value = cell_obj_1.value
                sheet.cell(row=j, column=9).font = Font(size=11, bold=True)
                cell_obj_2 = sheet_obj_5.cell(row=self.n, column=j)
                sheet.cell(row=j, column=10).value = cell_obj_2.value
                str = "{c1} has Expertise area at {c2}".format(c1=cell_obj_1.value, c2=cell_obj_2.value)
                print(str)
            print("---------------------------------------")
            wb.save(filepath)
    except Exception as e:
        print(e)

def main():
    try:
        PrintPs.print_ps()
        print("Enter The Valid Ps ID To Retrieve Data")
        n_1 = int(input())
        c_1 = CheckPs(n_1)
        val = c_1.check_ps()
        if val == 1:
            print("PS ID Present in Excel")
            E_1 = ExactPs(n_1)
            A = E_1.exactps()
            D_1 = Detailsps(A)
            D_1.detailsPs()
        else:
            print("Invalid PS ID")
    except Exception as e:
        print(e)

if __name__ == "__main__":
    main()